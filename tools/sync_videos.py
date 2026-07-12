import json
from pathlib import Path
import sys

import openpyxl


WORKBOOK_PATH = Path(r"C:\Users\123pr\Desktop\DSA LP VIDEOS FILE.xlsx")
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "data" / "videos.json"
SHEET_NAME = "GST VIDEOS"


def clean(value):
    return str(value).strip() if value is not None else ""


def main():
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else WORKBOOK_PATH
    existing_by_title = {}
    if OUTPUT_PATH.exists():
        try:
            existing_rows = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
            existing_by_title = {
                clean(row.get("title")): row
                for row in existing_rows
                if isinstance(row, dict) and clean(row.get("title"))
            }
        except json.JSONDecodeError:
            existing_by_title = {}

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    worksheet = workbook[SHEET_NAME]
    videos = []

    for row_index in range(6, worksheet.max_row + 1):
        title = clean(worksheet.cell(row_index, 2).value)
        loom_url = clean(worksheet.cell(row_index, 3).value)
        if not title and not loom_url:
            continue
        existing = existing_by_title.get(title, {})
        video = {
            "title": title,
            "loomUrl": loom_url,
            "sheet": SHEET_NAME,
            "row": row_index,
        }
        if existing.get("summary"):
            video["summary"] = existing["summary"]
        if existing.get("resources"):
            video["resources"] = existing["resources"]
        videos.append(video)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(videos, indent=2), encoding="utf-8")
    print(f"Synced {len(videos)} videos to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
